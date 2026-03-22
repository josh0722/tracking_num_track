from __future__ import annotations

import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook


BASE_URL = "https://apis.tracker.delivery"
SHEET2_NAME = "sheet2"
CARRIER_COL = 28  # AB
INVOICE_COL = 29  # AC
DELIVERY_STATUS_COL = 30  # AD
MOVEMENT_TIME_COL = 31  # AE
CURRENT_LOCATION_COL = 32  # AF
DELIVERY_CONTACT_COL = 33  # AG
PHONE_PATTERN = r"01[016789]-\d{3,4}-\d{4}"


class Sheet2SyncError(Exception):
    """Raised when Sheet2 sync flow fails."""


@dataclass
class TrackResult:
    status: str
    last_event_time: str
    last_event_location: str
    last_event_description: str
    delivery_contact: str


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9가-힣]", "", value.lower())


def _extract_tracking_numbers(value: str) -> list[str]:
    found = re.findall(r"\b\d{8,20}\b", value)
    seen: OrderedDict[str, None] = OrderedDict()
    for number in found:
        seen[number] = None
    return list(seen.keys())


def _format_event_time(value: str) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return value


def _extract_delivery_contact(description: str) -> str:
    if not description:
        return ""

    patterns = [
        rf"((?:배송담당|배송예정|담당사원|배송사원|배송기사|배송직원|집배원|담당자|담당직원|배달담당|기사님|기사)\s*:?\s*[^)]*?{PHONE_PATTERN})",
        rf"([가-힣A-Za-z]{{2,12}}\s+{PHONE_PATTERN})",
        rf"({PHONE_PATTERN})",
    ]
    for pattern in patterns:
        match = re.search(pattern, description)
        if match:
            return re.sub(r"\s+", " ", match.group(1)).strip()
    return ""


def _extract_delivery_contact_from_progresses(progresses: list[dict[str, Any]]) -> str:
    for progress in reversed(progresses):
        description = _normalize_text(progress.get("description"))
        contact = _extract_delivery_contact(description)
        if contact:
            return contact
    return ""


def _find_sheet2(workbook: Any) -> Any:
    for name in workbook.sheetnames:
        if name.lower() == SHEET2_NAME:
            return workbook[name]
    raise Sheet2SyncError("Sheet2 시트를 찾지 못했습니다.")


def _build_alias_map(known_ids: set[str]) -> dict[str, str]:
    aliases = {
        "cj대한통운": "kr.cjlogistics",
        "cjlogistics": "kr.cjlogistics",
        "대한통운": "kr.cjlogistics",
        "롯데택배": "kr.lotte",
        "롯데글로벌로지스": "kr.lotte",
        "lotte": "kr.lotte",
        "한진택배": "kr.hanjin",
        "한진": "kr.hanjin",
        "로젠택배": "kr.logen",
        "로젠": "kr.logen",
        "우체국택배": "kr.epost",
        "우체국": "kr.epost",
        "cu편의점택배": "kr.cupost",
        "cupost": "kr.cupost",
        "gspostbox": "kr.cvsnet",
        "gs포스트박스": "kr.cvsnet",
        "gs편의점택배": "kr.cvsnet",
        "경동택배": "kr.kdexp",
        "대신택배": "kr.daesin",
        "일양로지스": "kr.ilyanglogis",
        "합동택배": "kr.hdexp",
        "쿠팡로지스틱스": "kr.coupangls",
    }
    normalized: dict[str, str] = {}
    for key, carrier_id in aliases.items():
        if carrier_id in known_ids:
            normalized[_normalize_key(key)] = carrier_id
    return normalized


async def _fetch_carriers(client: httpx.AsyncClient) -> list[dict[str, Any]]:
    response = await client.get(f"{BASE_URL}/carriers", timeout=15.0)
    if response.status_code != 200:
        raise Sheet2SyncError("택배사 목록 조회에 실패했습니다.")
    payload = response.json()
    if not isinstance(payload, list):
        raise Sheet2SyncError("택배사 목록 형식이 올바르지 않습니다.")
    return payload


def _resolve_carrier_id(
    carrier_name: str,
    carrier_pairs: list[tuple[str, str]],
    alias_map: dict[str, str],
) -> str | None:
    target = _normalize_key(carrier_name)
    if not target:
        return None

    if target in alias_map:
        return alias_map[target]

    for normalized_name, carrier_id in carrier_pairs:
        if target == normalized_name:
            return carrier_id

    for normalized_name, carrier_id in carrier_pairs:
        if target in normalized_name or normalized_name in target:
            return carrier_id

    return None


async def _fetch_track_status(
    client: httpx.AsyncClient,
    carrier_id: str,
    tracking_number: str,
    cache: dict[tuple[str, str], TrackResult | None],
) -> TrackResult | None:
    key = (carrier_id, tracking_number)
    if key in cache:
        return cache[key]

    response = await client.get(
        f"{BASE_URL}/carriers/{carrier_id}/tracks/{tracking_number}",
        timeout=15.0,
    )
    if response.status_code != 200:
        cache[key] = None
        return None

    data = response.json()
    status = _normalize_text((data.get("state") or {}).get("text"))
    if not status:
        status = "상태 확인 중"

    progresses = data.get("progresses") or []
    last_progress = progresses[-1] if progresses else {}
    location = _normalize_text(((last_progress.get("location") or {}).get("name")))
    description = _normalize_text(last_progress.get("description"))
    event_time = _format_event_time(_normalize_text(last_progress.get("time")))
    delivery_contact = _extract_delivery_contact_from_progresses(progresses)

    track = TrackResult(
        status=status,
        last_event_time=event_time,
        last_event_location=location,
        last_event_description=description,
        delivery_contact=delivery_contact,
    )
    cache[key] = track
    return track


async def sync_sheet2_delivery_status(excel_path: str) -> dict[str, Any]:
    workbook_path = Path(excel_path).expanduser().resolve()
    if not workbook_path.exists():
        raise Sheet2SyncError(f"엑셀 파일을 찾지 못했습니다: {workbook_path}")

    workbook = load_workbook(workbook_path)
    sheet = _find_sheet2(workbook)

    summary: dict[str, Any] = {
        "updated_rows": 0,
        "missing_carrier_or_invoice": 0,
        "unresolved_carrier_rows": 0,
        "track_lookup_failed_rows": 0,
        "scanned_rows": max(sheet.max_row - 1, 0),
        "unresolved_carrier_samples": [],
    }

    unresolved_samples: list[dict[str, Any]] = []

    async with httpx.AsyncClient() as client:
        carriers = await _fetch_carriers(client)
        known_ids = {str(item.get("id")) for item in carriers if item.get("id")}
        alias_map = _build_alias_map(known_ids)

        carrier_pairs: list[tuple[str, str]] = []
        for carrier in carriers:
            carrier_id = _normalize_text(carrier.get("id"))
            name = _normalize_text(carrier.get("name"))
            if not carrier_id:
                continue
            carrier_pairs.append((_normalize_key(name), carrier_id))
            carrier_pairs.append((_normalize_key(carrier_id), carrier_id))

        track_cache: dict[tuple[str, str], TrackResult | None] = {}

        for row_idx in range(2, sheet.max_row + 1):
            carrier_name = _normalize_text(sheet.cell(row_idx, CARRIER_COL).value)
            invoice_raw = _normalize_text(sheet.cell(row_idx, INVOICE_COL).value)
            if not carrier_name or not invoice_raw:
                summary["missing_carrier_or_invoice"] += 1
                continue

            tracking_numbers = _extract_tracking_numbers(invoice_raw)
            if not tracking_numbers:
                summary["track_lookup_failed_rows"] += 1
                continue

            carrier_id = _resolve_carrier_id(carrier_name, carrier_pairs, alias_map)
            if not carrier_id:
                summary["unresolved_carrier_rows"] += 1
                if len(unresolved_samples) < 10:
                    unresolved_samples.append(
                        {
                            "row": row_idx,
                            "carrier": carrier_name,
                            "invoice": invoice_raw,
                        }
                    )
                continue

            primary_tracking_number = tracking_numbers[0]
            result = await _fetch_track_status(client, carrier_id, primary_tracking_number, track_cache)

            if result is None:
                summary["track_lookup_failed_rows"] += 1
                continue

            sheet.cell(row_idx, DELIVERY_STATUS_COL).value = result.status
            sheet.cell(row_idx, MOVEMENT_TIME_COL).value = result.last_event_time or None
            sheet.cell(row_idx, CURRENT_LOCATION_COL).value = result.last_event_location or None
            sheet.cell(row_idx, DELIVERY_CONTACT_COL).value = result.delivery_contact or None
            summary["updated_rows"] += 1

    summary["unresolved_carrier_samples"] = unresolved_samples
    workbook.save(workbook_path)

    return summary
