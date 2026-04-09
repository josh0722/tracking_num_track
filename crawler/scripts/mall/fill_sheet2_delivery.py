#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import threading
import time
from collections import OrderedDict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ORDER_NO_COL = 6  # F
USERNAME_COL = 23  # W
PASSWORD_COL = 24  # X
STATUS_COL = 12  # L
CARRIER_COL = 28  # AB
INVOICE_COL = 29  # AC
DELIVERY_STATUS_COL = 30  # AD
CANCEL_STYLE_TEMPLATE_ROW = 6
CANCEL_CLEAR_COLS = [
    13,  # M
    16,  # P
    17,  # Q
    18,  # R
    19,  # S
    25,  # Y
    26,  # Z
    27,  # AA
    28,  # AB
    29,  # AC
    30,  # AD
    31,  # AE
    32,  # AF
    33,  # AG
]


@dataclass
class TargetRow:
    row_idx: int
    order_no: str
    username: str
    password: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_blank(value: Any) -> bool:
    return normalize_text(value) == ""


def detect_npm_cmd(repo_root: Path) -> str:
    candidates: list[str] = []

    env_npm = os.environ.get("MALL_NPM_BIN", "").strip()
    if env_npm:
        candidates.append(env_npm)

    candidates.extend(
        [
            str(repo_root / "runtime" / "node" / ("npm.cmd" if os.name == "nt" else "npm")),
            str(repo_root / "runtime" / "node" / "node_modules" / "npm" / "bin" / "npm-cli.js"),
        ]
    )

    npm_from_path = shutil.which("npm")
    if npm_from_path:
        candidates.append(npm_from_path)

    candidates.extend(
        [
            "/opt/homebrew/bin/npm",
            "/usr/local/bin/npm",
            "/usr/bin/npm",
        ]
    )

    node_from_path = shutil.which("node")
    if node_from_path:
        candidates.append(str(Path(node_from_path).with_name("npm")))

    nvm_root = Path.home() / ".nvm" / "versions" / "node"
    if nvm_root.exists():
        for npm_bin in sorted(nvm_root.glob("*/bin/npm"), reverse=True):
            candidates.append(str(npm_bin))

    shell_candidates = [
        ["zsh", "-ic", "whence -p npm || command -v npm"],
        ["bash", "-ic", "command -v npm"],
    ]
    for shell_cmd in shell_candidates:
        try:
            proc = subprocess.run(
                shell_cmd,
                capture_output=True,
                text=True,
                timeout=6,
                check=False,
            )
            resolved = normalize_text(proc.stdout.splitlines()[0] if proc.stdout else "")
            if resolved:
                candidates.append(resolved)
        except Exception:
            continue

    seen: set[str] = set()
    for candidate in candidates:
        if not candidate:
            continue
        if candidate in seen:
            continue
        seen.add(candidate)
        path = Path(candidate).expanduser()
        if path.exists() and os.access(path, os.X_OK):
            return str(path)

    return ""


def detect_node_cmd(repo_root: Path) -> str:
    candidates: list[str] = []

    env_node = os.environ.get("MALL_NODE_BIN", "").strip()
    if env_node:
        candidates.append(env_node)

    candidates.extend(
        [
            str(repo_root / "runtime" / "node" / ("node.exe" if os.name == "nt" else "node")),
        ]
    )

    node_from_path = shutil.which("node")
    if node_from_path:
        candidates.append(node_from_path)

    candidates.extend(
        [
            "/opt/homebrew/bin/node",
            "/usr/local/bin/node",
            "/usr/bin/node",
        ]
    )

    if os.name == "nt":
        candidates.extend(
            [
                r"C:\Program Files\nodejs\node.exe",
                r"C:\Program Files (x86)\nodejs\node.exe",
            ]
        )

    seen: set[str] = set()
    for candidate in candidates:
        if not candidate:
            continue
        if candidate in seen:
            continue
        seen.add(candidate)
        path = Path(candidate).expanduser()
        if path.exists():
            return str(path)

    return ""


def normalize_order_no(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return re.sub(r"\D", "", f"{value}")

    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits or text


def find_sheet1(wb) -> Any:
    for name in wb.sheetnames:
        if name.lower() == "sheet1":
            return wb[name]
    raise ValueError("Sheet1 시트를 찾지 못했습니다.")


def build_sheet2(wb, sheet1, row_extra_tracking: dict[int, list[str]]) -> None:
    """
    Sheet2 생성: Sheet1의 모든 행 복사.
    row_extra_tracking = {sheet1_row_idx: [extra_tracking_2, extra_tracking_3, ...]}
    해당 행은 복사 후 추가 행을 이어붙여 송장번호별로 분리.
    """
    if "Sheet2" in wb.sheetnames:
        del wb["Sheet2"]
    sheet2 = wb.create_sheet("Sheet2")

    dst_row = 1
    for src_row in range(1, sheet1.max_row + 1):
        _copy_row(sheet1, src_row, sheet2, dst_row)
        dst_row += 1

        for extra_no in row_extra_tracking.get(src_row, []):
            _copy_row(sheet1, src_row, sheet2, dst_row)
            sheet2.cell(dst_row, INVOICE_COL).value = extra_no
            dst_row += 1


def _copy_row(src_sheet, src_row_idx: int, dst_sheet, dst_row_idx: int) -> None:
    for col_idx in range(1, src_sheet.max_column + 1):
        src_cell = src_sheet.cell(src_row_idx, col_idx)
        dst_cell = dst_sheet.cell(dst_row_idx, col_idx)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)


def collect_targets(sheet) -> tuple[list[TargetRow], int]:
    targets: list[TargetRow] = []
    skipped_prefilled_rows = 0
    for row_idx in range(2, sheet.max_row + 1):
        order_no = normalize_order_no(sheet.cell(row_idx, ORDER_NO_COL).value)
        username = normalize_text(sheet.cell(row_idx, USERNAME_COL).value)
        password = normalize_text(sheet.cell(row_idx, PASSWORD_COL).value)
        carrier = sheet.cell(row_idx, CARRIER_COL).value
        invoice = sheet.cell(row_idx, INVOICE_COL).value

        if not order_no or not username or not password:
            missing = []
            if not order_no:
                missing.append("order_no(F)")
            if not username:
                missing.append("username(W)")
            if not password:
                missing.append("password(X)")
            print(f"[SKIP] row {row_idx}: 필수 값 누락 ({', '.join(missing)})")
            continue

        # 요구사항: 택배사(AB), 송장번호(AC)가 모두 빈 행만 처리
        if not is_blank(carrier) or not is_blank(invoice):
            skipped_prefilled_rows += 1
            print(f"[SKIP] row {row_idx}: 이미 택배사/송장번호 입력됨 (carrier={carrier}, invoice={invoice})")
            continue

        targets.append(
            TargetRow(
                row_idx=row_idx,
                order_no=order_no,
                username=username,
                password=password,
            )
        )
    return targets, skipped_prefilled_rows


def build_accounts(targets: list[TargetRow]) -> tuple[list[dict[str, str]], dict[str, tuple[str, str]], list[str]]:
    unique_users: "OrderedDict[str, tuple[str, str]]" = OrderedDict()
    warnings: list[str] = []

    for target in targets:
        if target.username not in unique_users:
            account_id = f"excel-acct-{len(unique_users) + 1:03d}"
            unique_users[target.username] = (target.password, account_id)
            continue

        existing_password, _ = unique_users[target.username]
        if existing_password != target.password:
            warnings.append(
                f"동일 ID에 비밀번호가 여러 개입니다. 첫 번째 비밀번호를 사용합니다. id={target.username}"
            )

    accounts: list[dict[str, str]] = []
    account_id_to_cred: dict[str, tuple[str, str]] = {}
    for username, (password, account_id) in unique_users.items():
        accounts.append(
            {
                "accountId": account_id,
                "username": username,
                "password": password,
            }
        )
        account_id_to_cred[account_id] = (username, password)

    return accounts, account_id_to_cred, sorted(set(warnings))


def build_target_orders_by_username(targets: list[TargetRow]) -> dict[str, list[str]]:
    result: dict[str, set[str]] = {}
    for target in targets:
        if not target.username or not target.order_no:
            continue
        result.setdefault(target.username, set()).add(target.order_no)
    return {username: sorted(list(orders)) for username, orders in result.items()}


def parse_tracking_numbers(item: dict[str, Any]) -> set[str]:
    tracking_numbers: set[str] = set()

    for val in item.get("trackingNumbers", []) or []:
        number = normalize_order_no(val)
        if re.fullmatch(r"\d{8,20}", number):
            tracking_numbers.add(number)

    display_value = normalize_text(item.get("displayValue"))
    for number in re.findall(r"\b\d{8,20}\b", display_value):
        tracking_numbers.add(number)

    return tracking_numbers


def merge_scrape_results(
    rows: list[dict[str, Any]],
    account_id_to_cred: dict[str, tuple[str, str]],
) -> tuple[set[tuple[str, str]], dict[tuple[str, str], dict[str, Any]]]:
    canceled_keys: set[tuple[str, str]] = set()
    order_map: dict[tuple[str, str], dict[str, Any]] = {}

    for item in rows:
        account_username = normalize_text(item.get("accountUsername"))
        if not account_username:
            account_id = normalize_text(item.get("accountId"))
            account_username = account_id_to_cred.get(account_id, ("", ""))[0]

        order_no = normalize_order_no(item.get("orderNumber"))
        if not account_username or not order_no:
            continue

        key = (account_username, order_no)
        section = normalize_text(item.get("section"))

        if section == "cancel_status":
            canceled_keys.add(key)
            continue

        if section != "order_status":
            continue

        rec = order_map.setdefault(
            key,
            {
                "carrier": "",
                "tracking_numbers": set(),
            },
        )

        carrier = normalize_text(item.get("logisticsCompany"))
        if carrier and not rec["carrier"]:
            rec["carrier"] = carrier

        rec["tracking_numbers"].update(parse_tracking_numbers(item))

    return canceled_keys, order_map


def apply_cancel_style(sheet, target_row: int) -> None:
    for col in range(1, sheet.max_column + 1):
        source = sheet.cell(CANCEL_STYLE_TEMPLATE_ROW, col)
        dest = sheet.cell(target_row, col)
        dest._style = copy(source._style)

    sheet.cell(target_row, STATUS_COL).value = "취소완료"
    for col in CANCEL_CLEAR_COLS:
        sheet.cell(target_row, col).value = None


def update_workbook(
    workbook_path: Path,
    output_path: Path,
    canceled_keys: set[tuple[str, str]],
    order_map: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, Any]:
    wb = load_workbook(workbook_path)
    sheet = find_sheet1(wb)
    row_extra_tracking: dict[int, list[str]] = {}

    summary = {
        "target_rows": 0,
        "skipped_prefilled_rows": 0,
        "updated_rows": 0,
        "canceled_rows": 0,
        "not_found_rows": 0,
        "rows_with_multi_invoice": 0,
        "unmatched": [],
    }

    targets, skipped_prefilled_rows = collect_targets(sheet)
    summary["target_rows"] = len(targets)
    summary["skipped_prefilled_rows"] = skipped_prefilled_rows

    for target in targets:
        key = (target.username, target.order_no)

        if key in canceled_keys:
            apply_cancel_style(sheet, target.row_idx)
            summary["updated_rows"] += 1
            summary["canceled_rows"] += 1
            continue

        rec = order_map.get(key)
        if not rec:
            summary["not_found_rows"] += 1
            summary["unmatched"].append(
                {
                    "row": target.row_idx,
                    "username": target.username,
                    "order_no": target.order_no,
                    "reason": "scrape_result_not_found",
                }
            )
            continue

        carrier = normalize_text(rec.get("carrier"))
        tracking_numbers = sorted(rec.get("tracking_numbers", set()))

        if carrier:
            sheet.cell(target.row_idx, CARRIER_COL).value = carrier

        if tracking_numbers:
            invoice_value = tracking_numbers[0]
            sheet.cell(target.row_idx, INVOICE_COL).value = invoice_value
            if len(tracking_numbers) > 1:
                row_extra_tracking[target.row_idx] = tracking_numbers[1:]
                summary["rows_with_multi_invoice"] += 1

        summary["updated_rows"] += 1

    build_sheet2(wb, sheet, row_extra_tracking)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Sheet1 AB/AC(택배사/송장번호) 자동 반영")
    parser.add_argument(
        "--excel",
        default="",
        help="원본 엑셀 경로 (미입력 시 프로젝트 폴더의 '통합 문서1.xlsx' 우선 탐색)",
    )
    parser.add_argument(
        "--output",
        default="",
        help="결과 엑셀 경로 (미입력 시 원본 파일에 직접 반영)",
    )
    parser.add_argument(
        "--skip-crawl",
        action="store_true",
        help="이미 생성된 results-*.json만 사용(크롤링 생략)",
    )
    parser.add_argument(
        "--result-json",
        default="",
        help="--skip-crawl 사용 시 읽을 results-*.json 경로",
    )
    args = parser.parse_args()

    env_repo_root = os.environ.get("MALL_REPO_ROOT", "").strip()
    if env_repo_root:
        repo_root = Path(env_repo_root).expanduser().resolve()
    else:
        repo_root = Path(__file__).resolve().parents[2]

    npm_cmd = detect_npm_cmd(repo_root)
    node_cmd = detect_node_cmd(repo_root)

    tsx_bin = repo_root / "node_modules" / ".bin" / ("tsx.cmd" if os.name == "nt" else "tsx")
    tsx_cli = repo_root / "node_modules" / "tsx" / "dist" / "cli.mjs"
    if not tsx_bin.exists() and not tsx_cli.exists():
        raise RuntimeError(
            "필수 의존성(tsx)이 설치되지 않았습니다.\n"
            f"아래 명령을 먼저 실행해주세요:\ncd {repo_root} && npm install"
        )

    if not npm_cmd and not node_cmd:
        raise RuntimeError(
            "npm/node를 찾지 못했습니다. Node.js 설치 후 다시 실행해주세요.\n"
            "필요하면 환경변수로 경로를 지정할 수 있습니다.\n"
            "예: MALL_NPM_BIN=/opt/homebrew/bin/npm, MALL_NODE_BIN=/opt/homebrew/bin/node"
        )

    if args.excel:
        excel_path = Path(args.excel).expanduser().resolve()
    else:
        default_candidates = [
            repo_root / "통합 문서1.xlsx",
            Path.home() / "Downloads" / "통합 문서1.xlsx",
            Path.home() / "Downloads" / "통합 문서1.xlsx",
        ]
        excel_path = next((p for p in default_candidates if p.exists()), default_candidates[0]).resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾지 못했습니다: {excel_path}")

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        output_path = excel_path

    try:
        wb = load_workbook(excel_path)
        sheet = find_sheet1(wb)
        targets, skipped_prefilled_rows = collect_targets(sheet)
        wb.close()
    except Exception as exc:
        raise RuntimeError(
            f"엑셀 파일을 열 수 없습니다: {excel_path}\n"
            f"파일이 손상되었거나 다른 프로그램에서 사용 중일 수 있습니다.\n"
            f"상세: {exc}"
        ) from exc

    if not targets:
        if output_path != excel_path:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(excel_path, output_path)
        print(f"[DONE] 처리 대상 없음: AB/AC 빈 행이 없습니다. skipped_prefilled={skipped_prefilled_rows}")
        print(f"[DONE] output excel : {output_path}")
        raise SystemExit(0)

    print(f"[INFO] 처리 대상: {len(targets)}건, 이미 입력됨(skip): {skipped_prefilled_rows}건")
    accounts, account_id_to_cred, account_warnings = build_accounts(targets)
    target_orders_by_username = build_target_orders_by_username(targets)
    total_target_orders = sum(len(v) for v in target_orders_by_username.values())
    print(f"[INFO] 계정 수: {len(accounts)}, 타겟 주문 수: {total_target_orders}")
    for warning in account_warnings:
        print(f"[WARN] {warning}")

    run_stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = repo_root / "scripts" / "mall" / "output" / f"excel-run-{run_stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    accounts_path = run_dir / "accounts.generated.json"
    accounts_path.write_text(json.dumps(accounts, ensure_ascii=False, indent=2), encoding="utf-8")

    result_json_path: Path
    if not args.skip_crawl:
        env = os.environ.copy()
        env["MALL_ACCOUNTS_PATH"] = str(accounts_path)
        env["MALL_OUTPUT_DIR"] = str(run_dir)
        env["MALL_TARGET_ORDERS_JSON"] = json.dumps(target_orders_by_username, ensure_ascii=False)
        # 한글 경로/메시지가 stdout에 섞여도 cp949 디코딩으로 죽지 않도록 강제 UTF-8
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        path_segments = []
        if npm_cmd:
            path_segments.append(str(Path(npm_cmd).expanduser().parent))
        if node_cmd:
            path_segments.append(str(Path(node_cmd).expanduser().parent))
        if path_segments:
            env["PATH"] = f"{os.pathsep.join(path_segments)}{os.pathsep}{env.get('PATH', '')}"

        crawl_log_path = run_dir / "crawl-output.log"
        try:
            # node + tsx 직접 실행을 우선 사용 (번들 환경에서 npm.cmd는 불완전할 수 있음)
            if node_cmd and tsx_cli.exists():
                crawl_cmd = [node_cmd, str(tsx_cli), str(repo_root / "scripts" / "mall" / "scrapeMallShipments.ts")]
            elif npm_cmd:
                crawl_cmd = [npm_cmd, "run", "scrape:mall"]
            else:
                raise RuntimeError(
                    "크롤러를 실행할 수 없습니다. node+tsx 또는 npm이 필요합니다.\n"
                    f"node_cmd={node_cmd}, tsx_cli={tsx_cli}, npm_cmd={npm_cmd}"
                )

            print(f"[INFO] 크롤러 실행: {' '.join(crawl_cmd)}")
            print(f"[INFO] 크롤러 로그: {crawl_log_path}")

            creation_flags = 0
            if os.name == "nt":
                creation_flags = subprocess.CREATE_NO_WINDOW

            # 타임아웃 정책:
            #  - total_timeout: 계정 수에 비례 (계정당 90초, 최소 30분, 최대 4시간)
            #  - idle_timeout: stdout 무출력이 이 시간을 넘으면 hang으로 간주하고 종료
            total_timeout_s = max(1800, min(len(accounts) * 90, 14400))
            idle_timeout_s = 300  # 5분간 한 줄도 안 나오면 hang
            print(
                f"[INFO] 크롤러 타임아웃: total={total_timeout_s}s, idle={idle_timeout_s}s "
                f"(accounts={len(accounts)})"
            )

            proc = subprocess.Popen(
                crawl_cmd,
                cwd=repo_root,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=creation_flags,
            )

            last_output_at = time.monotonic()
            start_at = last_output_at
            output_lock = threading.Lock()
            reader_done = threading.Event()
            reader_error: list[BaseException] = []

            def _reader() -> None:
                nonlocal last_output_at
                try:
                    assert proc.stdout is not None
                    with open(crawl_log_path, "w", encoding="utf-8") as log_file:
                        for line in proc.stdout:
                            log_file.write(line)
                            log_file.flush()
                            print(line, end="")
                            with output_lock:
                                last_output_at = time.monotonic()
                except BaseException as e:  # noqa: BLE001
                    reader_error.append(e)
                finally:
                    reader_done.set()

            reader_thread = threading.Thread(target=_reader, daemon=True)
            reader_thread.start()

            kill_reason = ""
            while not reader_done.wait(timeout=5):
                now = time.monotonic()
                with output_lock:
                    idle = now - last_output_at
                elapsed = now - start_at
                if idle > idle_timeout_s:
                    kill_reason = (
                        f"무응답 {int(idle)}초 동안 출력 없음 "
                        f"(idle 임계 {idle_timeout_s}초)"
                    )
                    break
                if elapsed > total_timeout_s:
                    kill_reason = (
                        f"총 실행 {int(elapsed)}초 초과 "
                        f"(총 임계 {total_timeout_s}초, accounts={len(accounts)})"
                    )
                    break

            if kill_reason:
                try:
                    proc.kill()
                except Exception:
                    pass
                reader_thread.join(timeout=10)
                raise RuntimeError(f"크롤링 강제 종료: {kill_reason}")

            reader_thread.join(timeout=10)
            return_code = proc.wait(timeout=30)
            if reader_error:
                raise RuntimeError(f"크롤러 stdout 읽기 실패: {reader_error[0]}")
            if return_code != 0:
                raise subprocess.CalledProcessError(return_code, crawl_cmd)
        except subprocess.TimeoutExpired:
            if proc is not None:
                proc.kill()
            raise RuntimeError(
                "크롤링 프로세스 정리 중 시간 초과: 프로세스가 응답하지 않아 종료했습니다."
            )
        except subprocess.CalledProcessError as exc:
            # 로그 파일 내용을 에러 메시지에 포함
            log_tail = ""
            if crawl_log_path.exists():
                log_content = crawl_log_path.read_text(encoding="utf-8", errors="replace")
                lines = log_content.strip().splitlines()
                log_tail = "\n".join(lines[-20:]) if lines else "(빈 로그)"
            if exc.returncode == 127:
                raise RuntimeError(
                    "크롤링 실행 도구를 찾지 못했습니다. 의존성 설치 후 다시 시도해주세요.\n"
                    f"cd {repo_root} && npm install"
                ) from exc
            raise RuntimeError(
                f"크롤링 실행 실패 (exit code: {exc.returncode})\n"
                f"로그 파일: {crawl_log_path}\n"
                f"--- 마지막 로그 ---\n{log_tail}"
            ) from exc
        result_jsons = sorted(run_dir.glob("results-*.json"), key=lambda p: p.stat().st_mtime)
        if not result_jsons:
            raise RuntimeError(f"크롤링 결과 JSON을 찾지 못했습니다: {run_dir}")
        result_json_path = result_jsons[-1]
    else:
        if not args.result_json:
            raise RuntimeError("--skip-crawl 사용 시 --result-json 경로를 함께 지정해야 합니다.")
        result_json_path = Path(args.result_json).expanduser().resolve()
        if not result_json_path.exists():
            raise FileNotFoundError(f"results json을 찾지 못했습니다: {result_json_path}")

    try:
        scraped_rows = json.loads(result_json_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise RuntimeError(
            f"크롤링 결과 JSON 파싱 실패: {result_json_path}\n"
            f"파일이 손상되었을 수 있습니다. 다시 실행해주세요.\n"
            f"상세: {exc}"
        ) from exc

    canceled_keys, order_map = merge_scrape_results(scraped_rows, account_id_to_cred)

    backup_path = ""
    if not args.output:
        output_path = excel_path
        backup = excel_path.with_name(f"{excel_path.stem}_backup_{run_stamp}{excel_path.suffix}")
        shutil.copy2(excel_path, backup)
        backup_path = str(backup)

    summary = update_workbook(
        workbook_path=excel_path,
        output_path=output_path,
        canceled_keys=canceled_keys,
        order_map=order_map,
    )

    summary_path = run_dir / "excel-update-summary.json"
    summary_payload = {
        "excel_input": str(excel_path),
        "excel_output": str(output_path),
        "excel_backup": backup_path,
        "scrape_result": str(result_json_path),
        "accounts_count": len(accounts),
        "account_warnings": account_warnings,
        "targets_count": len(targets),
        "skipped_prefilled_rows": skipped_prefilled_rows,
        **summary,
    }
    summary_path.write_text(json.dumps(summary_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[DONE] output excel : {output_path}")
    if backup_path:
        print(f"[DONE] backup excel : {backup_path}")
    print(f"[DONE] scrape json  : {result_json_path}")
    print(f"[DONE] summary json : {summary_path}")
    print(
        "[DONE] updated={updated_rows}, canceled={canceled_rows}, not_found={not_found_rows}, multi_invoice={rows_with_multi_invoice}".format(
            **summary
        )
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"[ERROR] {exc}")
        raise SystemExit(1)
