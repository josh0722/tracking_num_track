#!/usr/bin/env python3
"""SK스토아 적립금/멤버십할인 조회 워커.

- 입력: `SK스토아 적립금 조회.xlsx` (또는 동일 구조의 엑셀)
  - `데이터` 시트: A=순번, B=아이디, C=패스워드, D=이름
- 출력: `완료` 시트를 새로 생성 (A~D + E=소멸예정일, F=금액, G=멤버십할인, H=비고)
  - B 또는 C 가 비었거나, B/C 셀 fill 이 순수 초록(`FF00FF00`) 이면 해당 계정은
    로그인 시도 없이 E/F/G = '조회불가'
  - 계정에 소멸예정 적립금이 N(>=1)개이면 N행으로 펼치고 A/B/C/D/G/H 는
    vertical merge 로 한 번만 보이게 한다.

실행 원리는 `fill_sheet2_delivery.py` 와 동일하다. 계정 리스트를 뽑아
`accounts.generated.json` 을 생성하고, Node + tsx 로 `scrapeMallSavings.ts` 를
스폰한 뒤, 결과 JSON (`savings-*.json`) 을 읽어 `완료` 시트를 새로 만든다.
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# fill_sheet2_delivery.py 의 node/npm 탐색 로직을 재사용한다.
from fill_sheet2_delivery import detect_node_cmd, detect_npm_cmd  # type: ignore

SEQ_COL = 1  # A
USERNAME_COL = 2  # B
PASSWORD_COL = 3  # C
NAME_COL = 4  # D
EXPIRE_DATE_COL = 5  # E
EXPIRE_AMOUNT_COL = 6  # F
MEMBERSHIP_COL = 7  # G
NOTE_COL = 8  # H

# 스킵 규칙: 순수 초록
GREEN_SKIP_FILLS = {"FF00FF00", "00FF00"}

COMPLETE_SHEET_NAME = "완료"
DATA_SHEET_NAME = "데이터"

HEADER = ["순번", "아이디", "패스워드", "이름", "소멸예정일", "금액", "멤버십할인", "비고"]
MERGE_COLS_MULTI = [SEQ_COL, USERNAME_COL, PASSWORD_COL, NAME_COL, MEMBERSHIP_COL, NOTE_COL]

HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass
class Entry:
    seq: Any
    username: str
    password: str
    name: str
    skip: bool
    skip_reason: str
    account_id: str  # skip 이면 ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _fill_hex(cell) -> str:
    try:
        fill = cell.fill
        if not fill or fill.patternType != "solid":
            return ""
        fg = fill.fgColor
        if fg is None:
            return ""
        if getattr(fg, "type", None) == "rgb":
            return (fg.value or "").upper()
        return ""
    except Exception:
        return ""


def is_skip_row(username_cell, password_cell) -> tuple[bool, str]:
    username = normalize_text(username_cell.value)
    password = normalize_text(password_cell.value)
    if not username or not password:
        return True, "아이디/패스워드 비어있음"
    for cell in (username_cell, password_cell):
        hexval = _fill_hex(cell)
        # fgColor.value 가 'FF00FF00' 또는 '00FF00' 양쪽 형태로 올 수 있다.
        if hexval in GREEN_SKIP_FILLS or hexval.endswith("00FF00"):
            return True, "초록색 표시"
    return False, ""


def find_data_sheet(wb) -> Any:
    for name in wb.sheetnames:
        if name == DATA_SHEET_NAME:
            return wb[name]
    # fallback: 첫 시트
    for name in wb.sheetnames:
        if name != COMPLETE_SHEET_NAME:
            return wb[name]
    raise ValueError(f"'{DATA_SHEET_NAME}' 시트를 찾을 수 없습니다.")


def collect_entries(ws) -> list[Entry]:
    entries: list[Entry] = []
    acct_counter = 0
    for row_idx in range(2, ws.max_row + 1):
        seq = ws.cell(row_idx, SEQ_COL).value
        username_cell = ws.cell(row_idx, USERNAME_COL)
        password_cell = ws.cell(row_idx, PASSWORD_COL)
        name = normalize_text(ws.cell(row_idx, NAME_COL).value)
        # 전부 빈 행은 무시
        if seq is None and not normalize_text(username_cell.value) and not normalize_text(password_cell.value) and not name:
            continue
        skip, reason = is_skip_row(username_cell, password_cell)
        account_id = ""
        if not skip:
            acct_counter += 1
            account_id = f"savings-acct-{acct_counter:03d}"
        entries.append(
            Entry(
                seq=seq,
                username=normalize_text(username_cell.value),
                password=normalize_text(password_cell.value),
                name=name,
                skip=skip,
                skip_reason=reason,
                account_id=account_id,
            )
        )
    return entries


def build_accounts_json(entries: list[Entry]) -> list[dict[str, str]]:
    accounts: list[dict[str, str]] = []
    for entry in entries:
        if entry.skip:
            continue
        accounts.append(
            {
                "accountId": entry.account_id,
                "username": entry.username,
                "password": entry.password,
            }
        )
    return accounts


def run_crawler(
    repo_root: Path,
    accounts_path: Path,
    run_dir: Path,
    accounts_count: int,
) -> Path:
    npm_cmd = detect_npm_cmd(repo_root)
    node_cmd = detect_node_cmd(repo_root)

    tsx_bin = repo_root / "node_modules" / ".bin" / ("tsx.cmd" if os.name == "nt" else "tsx")
    tsx_cli = repo_root / "node_modules" / "tsx" / "dist" / "cli.mjs"
    if not tsx_bin.exists() and not tsx_cli.exists():
        raise RuntimeError(
            "필수 의존성(tsx)이 설치되지 않았습니다.\n"
            f"아래 명령을 먼저 실행해주세요:\ncd {repo_root} && npm install"
        )

    if not npm_cmd and not node_cmd:
        raise RuntimeError(
            "npm/node를 찾지 못했습니다. Node.js 설치 후 다시 실행해주세요.\n"
            "필요하면 환경변수로 경로를 지정할 수 있습니다.\n"
            "예: MALL_NPM_BIN=/opt/homebrew/bin/npm, MALL_NODE_BIN=/opt/homebrew/bin/node"
        )

    env = os.environ.copy()
    env["MALL_ACCOUNTS_PATH"] = str(accounts_path)
    env["MALL_OUTPUT_DIR"] = str(run_dir)
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    path_segments = []
    if npm_cmd:
        path_segments.append(str(Path(npm_cmd).expanduser().parent))
    if node_cmd:
        path_segments.append(str(Path(node_cmd).expanduser().parent))
    if path_segments:
        env["PATH"] = f"{os.pathsep.join(path_segments)}{os.pathsep}{env.get('PATH', '')}"

    if node_cmd and tsx_cli.exists():
        crawl_cmd = [node_cmd, str(tsx_cli), str(repo_root / "scripts" / "mall" / "scrapeMallSavings.ts")]
    elif npm_cmd:
        # package.json 에 scrape:savings 스크립트가 없다면 node+tsx 직접 실행으로 떨어진다.
        crawl_cmd = [npm_cmd, "run", "scrape:mall:savings"]
    else:
        raise RuntimeError("크롤러를 실행할 수 없습니다.")

    crawl_log_path = run_dir / "crawl-output.log"
    print(f"[INFO] 크롤러 실행: {' '.join(crawl_cmd)}")
    print(f"[INFO] 크롤러 로그: {crawl_log_path}")

    creation_flags = 0
    if os.name == "nt":
        creation_flags = subprocess.CREATE_NO_WINDOW  # type: ignore[attr-defined]

    total_timeout_s = max(1800, min(accounts_count * 90, 14400))
    idle_timeout_s = 300
    print(
        f"[INFO] 크롤러 타임아웃: total={total_timeout_s}s, idle={idle_timeout_s}s "
        f"(accounts={accounts_count})"
    )

    proc = subprocess.Popen(
        crawl_cmd,
        cwd=repo_root,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=creation_flags,
    )

    last_output_at = time.monotonic()
    start_at = last_output_at
    output_lock = threading.Lock()
    reader_done = threading.Event()
    reader_error: list[BaseException] = []

    def _reader() -> None:
        nonlocal last_output_at
        try:
            assert proc.stdout is not None
            with open(crawl_log_path, "w", encoding="utf-8") as log_file:
                for line in proc.stdout:
                    log_file.write(line)
                    log_file.flush()
                    print(line, end="")
                    with output_lock:
                        last_output_at = time.monotonic()
        except BaseException as e:  # noqa: BLE001
            reader_error.append(e)
        finally:
            reader_done.set()

    reader_thread = threading.Thread(target=_reader, daemon=True)
    reader_thread.start()

    kill_reason = ""
    while not reader_done.wait(timeout=5):
        now = time.monotonic()
        with output_lock:
            idle = now - last_output_at
        elapsed = now - start_at
        if idle > idle_timeout_s:
            kill_reason = f"무응답 {int(idle)}초 동안 출력 없음 (idle 임계 {idle_timeout_s}초)"
            break
        if elapsed > total_timeout_s:
            kill_reason = (
                f"총 실행 {int(elapsed)}초 초과 "
                f"(총 임계 {total_timeout_s}초, accounts={accounts_count})"
            )
            break

    if kill_reason:
        try:
            proc.kill()
        except Exception:
            pass
        reader_thread.join(timeout=10)
        raise RuntimeError(f"크롤링 강제 종료: {kill_reason}")

    reader_thread.join(timeout=10)
    return_code = proc.wait(timeout=30)
    if reader_error:
        raise RuntimeError(f"크롤러 stdout 읽기 실패: {reader_error[0]}")
    if return_code != 0:
        log_tail = ""
        if crawl_log_path.exists():
            log_content = crawl_log_path.read_text(encoding="utf-8", errors="replace")
            lines = log_content.strip().splitlines()
            log_tail = "\n".join(lines[-20:]) if lines else "(빈 로그)"
        raise RuntimeError(
            f"크롤링 실행 실패 (exit code: {return_code})\n"
            f"로그 파일: {crawl_log_path}\n"
            f"--- 마지막 로그 ---\n{log_tail}"
        )

    result_jsons = sorted(run_dir.glob("savings-*.json"), key=lambda p: p.stat().st_mtime)
    if not result_jsons:
        raise RuntimeError(f"크롤링 결과 JSON을 찾지 못했습니다: {run_dir}")
    return result_jsons[-1]


def build_complete_sheet(wb, entries: list[Entry], results_by_username: dict[str, dict[str, Any]]) -> dict[str, int]:
    if COMPLETE_SHEET_NAME in wb.sheetnames:
        del wb[COMPLETE_SHEET_NAME]
    ws = wb.create_sheet(COMPLETE_SHEET_NAME)

    # 헤더
    for idx, title in enumerate(HEADER, start=1):
        c = ws.cell(1, idx, title)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    summary = {
        "total_entries": len(entries),
        "skipped_entries": 0,
        "error_entries": 0,
        "ok_entries": 0,
        "multi_expiration_entries": 0,
        "total_rows_written": 0,
    }

    current_row = 2
    for entry in entries:
        if entry.skip:
            summary["skipped_entries"] += 1
            ws.cell(current_row, SEQ_COL, entry.seq)
            ws.cell(current_row, USERNAME_COL, entry.username or None)
            ws.cell(current_row, PASSWORD_COL, entry.password or None)
            ws.cell(current_row, NAME_COL, entry.name or None)
            ws.cell(current_row, EXPIRE_DATE_COL, "조회불가")
            ws.cell(current_row, EXPIRE_AMOUNT_COL, "조회불가")
            ws.cell(current_row, MEMBERSHIP_COL, "조회불가")
            ws.cell(current_row, NOTE_COL, None)
            current_row += 1
            summary["total_rows_written"] += 1
            continue

        row_data = results_by_username.get(entry.username)
        if not row_data or row_data.get("error"):
            summary["error_entries"] += 1
            ws.cell(current_row, SEQ_COL, entry.seq)
            ws.cell(current_row, USERNAME_COL, entry.username)
            ws.cell(current_row, PASSWORD_COL, entry.password)
            ws.cell(current_row, NAME_COL, entry.name or None)
            ws.cell(current_row, EXPIRE_DATE_COL, "조회불가")
            ws.cell(current_row, EXPIRE_AMOUNT_COL, "조회불가")
            ws.cell(current_row, MEMBERSHIP_COL, "조회불가")
            ws.cell(current_row, NOTE_COL, None)
            current_row += 1
            summary["total_rows_written"] += 1
            continue

        membership = row_data.get("membershipDiscount")
        expirations = row_data.get("expirations") or []
        summary["ok_entries"] += 1

        if not expirations:
            ws.cell(current_row, SEQ_COL, entry.seq)
            ws.cell(current_row, USERNAME_COL, entry.username)
            ws.cell(current_row, PASSWORD_COL, entry.password)
            ws.cell(current_row, NAME_COL, entry.name or None)
            ws.cell(current_row, EXPIRE_DATE_COL, None)
            ws.cell(current_row, EXPIRE_AMOUNT_COL, None)
            ws.cell(current_row, MEMBERSHIP_COL, membership)
            ws.cell(current_row, NOTE_COL, None)
            current_row += 1
            summary["total_rows_written"] += 1
            continue

        n = len(expirations)
        if n > 1:
            summary["multi_expiration_entries"] += 1
        start_row = current_row
        for i, exp in enumerate(expirations):
            if i == 0:
                ws.cell(current_row, SEQ_COL, entry.seq)
                ws.cell(current_row, USERNAME_COL, entry.username)
                ws.cell(current_row, PASSWORD_COL, entry.password)
                ws.cell(current_row, NAME_COL, entry.name or None)
                ws.cell(current_row, MEMBERSHIP_COL, membership)
                ws.cell(current_row, NOTE_COL, None)
            ws.cell(current_row, EXPIRE_DATE_COL, exp.get("expireDate") or None)
            ws.cell(current_row, EXPIRE_AMOUNT_COL, exp.get("amount"))
            current_row += 1
            summary["total_rows_written"] += 1

        if n > 1:
            end_row = start_row + n - 1
            for col in MERGE_COLS_MULTI:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col,
                )
            # 병합 셀 중앙 정렬
            for col in MERGE_COLS_MULTI:
                ws.cell(start_row, col).alignment = CENTER

    # 열 너비 대충 맞추기
    widths = {1: 6, 2: 26, 3: 18, 4: 10, 5: 14, 6: 12, 7: 12, 8: 14}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="SK스토아 적립금/멤버십할인 조회")
    parser.add_argument("--excel", default="", help="원본 엑셀 경로")
    parser.add_argument("--output", default="", help="결과 엑셀 경로 (미지정 시 원본에 직접 반영 + 백업 생성)")
    parser.add_argument("--skip-crawl", action="store_true", help="크롤링 생략 (--result-json 필요)")
    parser.add_argument("--result-json", default="", help="--skip-crawl 시 사용할 savings-*.json 경로")
    args = parser.parse_args()

    env_repo_root = os.environ.get("MALL_REPO_ROOT", "").strip()
    if env_repo_root:
        repo_root = Path(env_repo_root).expanduser().resolve()
    else:
        repo_root = Path(__file__).resolve().parents[2]

    if args.excel:
        excel_path = Path(args.excel).expanduser().resolve()
    else:
        default_candidates = [
            repo_root.parent / "SK스토아 적립금 조회.xlsx",
            repo_root / "SK스토아 적립금 조회.xlsx",
            Path.home() / "Downloads" / "SK스토아 적립금 조회.xlsx",
        ]
        excel_path = next((p for p in default_candidates if p.exists()), default_candidates[0]).resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾지 못했습니다: {excel_path}")

    output_path = Path(args.output).expanduser().resolve() if args.output else excel_path

    try:
        wb_probe = load_workbook(excel_path)
        ws_data = find_data_sheet(wb_probe)
        entries = collect_entries(ws_data)
        wb_probe.close()
    except Exception as exc:
        raise RuntimeError(
            f"엑셀 파일을 열 수 없습니다: {excel_path}\n"
            f"파일이 손상되었거나 다른 프로그램에서 사용 중일 수 있습니다.\n"
            f"상세: {exc}"
        ) from exc

    if not entries:
        raise RuntimeError(f"'{DATA_SHEET_NAME}' 시트에서 처리할 행을 찾지 못했습니다.")

    accounts = build_accounts_json(entries)
    skipped_count = sum(1 for e in entries if e.skip)
    print(f"[INFO] 총 {len(entries)}행, 크롤 대상 {len(accounts)}행, 스킵 {skipped_count}행")

    run_stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = repo_root / "scripts" / "mall" / "output" / f"savings-run-{run_stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    accounts_path = run_dir / "accounts.generated.json"
    accounts_path.write_text(json.dumps(accounts, ensure_ascii=False, indent=2), encoding="utf-8")

    results_by_username: dict[str, dict[str, Any]] = {}

    if accounts:
        if not args.skip_crawl:
            result_json_path = run_crawler(repo_root, accounts_path, run_dir, len(accounts))
        else:
            if not args.result_json:
                raise RuntimeError("--skip-crawl 사용 시 --result-json 경로를 함께 지정해야 합니다.")
            result_json_path = Path(args.result_json).expanduser().resolve()
            if not result_json_path.exists():
                raise FileNotFoundError(f"savings json을 찾지 못했습니다: {result_json_path}")

        try:
            rows = json.loads(result_json_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise RuntimeError(
                f"크롤링 결과 JSON 파싱 실패: {result_json_path}\n상세: {exc}"
            ) from exc

        for row in rows:
            username = normalize_text(row.get("accountUsername"))
            if username:
                results_by_username[username] = row
    else:
        print("[INFO] 크롤 대상 계정이 없어 크롤링을 생략합니다.")
        result_json_path = None  # type: ignore[assignment]

    # 백업 생성 (output 미지정 시만)
    backup_path = ""
    if not args.output:
        backup = excel_path.with_name(f"{excel_path.stem}_backup_{run_stamp}{excel_path.suffix}")
        shutil.copy2(excel_path, backup)
        backup_path = str(backup)

    wb = load_workbook(excel_path)
    summary = build_complete_sheet(wb, entries, results_by_username)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    summary_path = run_dir / "savings-update-summary.json"
    summary_payload = {
        "excel_input": str(excel_path),
        "excel_output": str(output_path),
        "excel_backup": backup_path,
        "scrape_result": str(result_json_path) if result_json_path else "",
        "accounts_count": len(accounts),
        **summary,
    }
    summary_path.write_text(json.dumps(summary_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[DONE] output excel : {output_path}")
    if backup_path:
        print(f"[DONE] backup excel : {backup_path}")
    if result_json_path:
        print(f"[DONE] scrape json  : {result_json_path}")
    print(f"[DONE] summary json : {summary_path}")
    print(
        "[DONE] total={total_entries}, ok={ok_entries}, error={error_entries}, "
        "skipped={skipped_entries}, multi={multi_expiration_entries}, rows={total_rows_written}".format(
            **summary
        )
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"[ERROR] {exc}")
        raise SystemExit(1)
